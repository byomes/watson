"""jobs/documents/excel.py — Read and create Excel .xlsx files."""
import logging
import re

log = logging.getLogger(__name__)


def read_xlsx(path: str) -> str:
    import openpyxl
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        lines = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            lines.append(f"Sheet: {sheet}")
            for row in ws.iter_rows(values_only=True):
                row_text = "\t".join("" if v is None else str(v) for v in row)
                if row_text.strip():
                    lines.append(row_text)
        return "\n".join(lines)
    except Exception as exc:
        log.error("read_xlsx failed: %s", exc)
        return f"Error reading Excel file: {exc}"


def create_xlsx(data: list, path: str, headers: list = None) -> bool:
    import openpyxl
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        if headers:
            ws.append(headers)
        for row in data:
            ws.append(row)
        wb.save(path)
        return True
    except Exception as exc:
        log.error("create_xlsx failed: %s", exc)
        return False


def run(message: str = None) -> str:
    if not message:
        return "Excel skill ready."
    match = re.search(r'[\w/~.-]+\.xlsx?', message, re.IGNORECASE)
    if not match:
        return "Excel skill ready. Provide a file path to read a spreadsheet."
    path = match.group(0).replace("~", __import__("os").path.expanduser("~"))
    text = read_xlsx(path)
    if not text:
        return f"No data extracted from {path}."
    preview = text[:1000]
    suffix = f"\n\n[{len(text)} chars total]" if len(text) > 1000 else ""
    return f"Excel: {path}\n\n{preview}{suffix}"
